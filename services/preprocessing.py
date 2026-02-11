import os
import re
import io
import time
import zipfile
import tempfile
from datetime import datetime, timezone, timedelta
from io import BytesIO
import xml.etree.ElementTree as ET
import pandas as pd
import requests
import pandas as pd
from io import BytesIO
from typing import Dict, Tuple, Optional

def process_file_to_dataframe( file_bytes: BytesIO, file_name: str, sheet_name: Optional[str] = None, 
                              header: int = 0, csv_delimiter: str = "comma") -> pd.DataFrame:
    # Your existing read logic here
    if file_name.lower().endswith(".xlsx") or file_name.lower().endswith(".xls"):
        from services.preprocessing import read_excel_with_repair
        return read_excel_with_repair(file_bytes, sheet_name or 0, header, file_name)
    
    elif file_name.lower().endswith(".csv"):
        # CSV logic
        text = file_bytes.getvalue().decode("utf-8", errors="ignore")
        delimiter_map = {
            "comma": ",", "semicolon": ";", 
            "tab": "\t", "pipe": "|"
        }
        delimiter = delimiter_map.get(csv_delimiter, ",")
        
        import io
        return pd.read_csv(io.StringIO(text), delimiter=delimiter, header=header)
    
    else:
        raise ValueError(f"Unsupported file type: {file_name}")

def extract_columns_metadata(df: pd.DataFrame) -> Dict:
    """
    Extract column information from DataFrame
    Returns: {
        'column_name': {
            'dtype': str,
            'null_count': int,
            'null_percentage': float,
            'unique_count': int,
            'sample_values': list,
            'inferred_type': str  # 'numeric', 'text', 'date', 'boolean'
        }
    }
    """
    columns_info = {}
    
    for col in df.columns:
        col_data = df[col]
        
        # Infer semantic type
        inferred_type = infer_column_type(col_data)
        
        columns_info[col] = {
            'dtype': str(col_data.dtype),
            'null_count': int(col_data.isnull().sum()),
            'null_percentage': round(col_data.isnull().sum() / len(df) * 100, 2),
            'unique_count': int(col_data.nunique()),
            'sample_values': col_data.dropna().head(5).astype(str).tolist(),
            'inferred_type': inferred_type
        }
    
    return columns_info

def infer_column_type(series: pd.Series) -> str:
    """Infer semantic type of column"""
    # Numeric
    if pd.api.types.is_numeric_dtype(series):
        if series.dropna().apply(lambda x: x == int(x)).all():
            return 'integer'
        return 'float'
    
    # Boolean
    if pd.api.types.is_bool_dtype(series):
        return 'boolean'
    
    # Date
    try:
        pd.to_datetime(series.dropna(), errors='raise')
        return 'date'
    except:
        pass
    
    # Default to text
    return 'text'

def backup_file_in_sharepoint(file_id, file_name, parent_folder_id, headers, SITE_ID, DRIVE_ID, backup_folder_path):
    """
    Create a backup copy of a file in SharePoint before reading it.
    Backup file will be named: original_name_YYYYMMDD_HHMMSS_Backup.ext
    
    Returns:
    - dict with backup info if successful
    - Raises Exception if backup fails (stops the process)
    """
    try:
        # Get the backup folder ID (where backup will be saved)
        backup_parent_id = get_parent_folder_id(backup_folder_path, headers, SITE_ID, DRIVE_ID)
        
        if not backup_parent_id:
            raise Exception(f"Backup folder not found: {backup_folder_path}")
        
        # Generate backup name with timestamp (include time to avoid duplicates)
        timestamp = (datetime.now(timezone.utc) + timedelta(hours=7)).strftime("%Y%m%d_%H%M%S")
        name_parts = file_name.rsplit('.', 1)
        if len(name_parts) == 2:
            backup_name = f"{name_parts[0]}_{timestamp}_Backup.{name_parts[1]}"
        else:
            backup_name = f"{file_name}_{timestamp}_Backup"
        
        print(f"  → Creating backup: {backup_name}")
        print(f"  → Backup location: {backup_folder_path}")
        
        copy_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{file_id}/copy"
        copy_body = {
            "parentReference": {"id": backup_parent_id},  # Changed from parent_folder_id
            "name": backup_name
        }
        
        response = requests.post(copy_url, headers=headers, json=copy_body)
        
        if response.status_code != 202:
            # Backup failed - STOP PROCESS
            error_msg = response.json().get("error", {}).get("message", response.text)
            raise Exception(f"Backup creation failed (HTTP {response.status_code}): {error_msg}")
        
        # Step 3: Wait and verify backup exists in BACKUP folder
        print(f"  → Waiting for backup to complete...")
        time.sleep(3)  # Quick wait for small files
        
        # Verify backup exists in the backup folder
        verify_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{backup_parent_id}/children"
        
        for attempt in range(10):  # Max 10 attempts = 30 seconds
            verify_response = requests.get(verify_url, headers=headers)
            
            if verify_response.status_code == 200:
                items = verify_response.json().get("value", [])
                backup_exists = any(item.get("name") == backup_name for item in items)
                
                if backup_exists:
                    print(f"  ✓ Backup confirmed: {backup_name}")
                    return {
                        "success": True,
                        "backup_name": backup_name
                    }
            
            if attempt < 9:
                time.sleep(3)
        
        # If we reach here, backup not verified - STOP PROCESS
        raise Exception(f"Backup '{backup_name}' not found after 30 seconds - process stopped")
            
    except Exception as e:
        # Re-raise exception to stop the entire process
        error_msg = f"BACKUP FAILED: {str(e)}"
        print(f"  ✗ {error_msg}")
        raise Exception(error_msg)

def get_parent_folder_id(folder_path, headers, SITE_ID, DRIVE_ID):
    """
    Get the folder ID from folder path for backup location
    """
    try:
        folder_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/root:/{folder_path}"
        response = requests.get(folder_url, headers=headers)
        
        if response.status_code == 200:
            return response.json().get("id")
        else:
            print(f"  ⚠ Could not get parent folder ID: {response.status_code}")
            return None
    except Exception as e:
        print(f"  ⚠ Error getting parent folder ID: {e}")
        return None


def list_all_files(folder_url, folder_path, headers, SITE_ID, DRIVE_ID, parent_folder=None):
    """Recursively collect all files from a folder and its subfolders."""
    result = []

    r = requests.get(folder_url, headers=headers)
    if r.status_code != 200:
        print("Failed to read folder:", folder_url)
        return result

    items = r.json().get("value", [])

    for item in items:

        # If it's a file
        if "file" in item:
            result.append({
                "name": item["name"],
                "download_url": item["@microsoft.graph.downloadUrl"],
                "file_id": item["id"],
                "folder_name": parent_folder,
                "parent_folder_id": item.get("parentReference", {}).get("id")
            })

        # If it's a subfolder → recurse using PATH
        if "folder" in item:
            sub_path = folder_path + "/" + item["name"]
            sub_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/root:/{sub_path}:/children"

            # extend recursively
            result.extend(
                list_all_files(sub_url, sub_path, headers, SITE_ID, DRIVE_ID, parent_folder=item["name"])
            )

    return result

def repair_excel_styles(file_bytes):
    """
    Repair corrupt Excel file dengan menghapus/fix stylesheet yang bermasalah
    """
    try:
        # Excel adalah ZIP file, extract dan repair
        with zipfile.ZipFile(file_bytes, 'r') as zip_ref:
            # Baca semua files
            file_list = zip_ref.namelist()
            
            # Create new ZIP in memory
            repaired_bytes = BytesIO()
            with zipfile.ZipFile(repaired_bytes, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                
                for filename in file_list:
                    file_data = zip_ref.read(filename)
                    
                    # Skip atau fix styles.xml yang bermasalah
                    if filename == 'xl/styles.xml':
                        try:
                            # Parse dan clean styles
                            root = ET.fromstring(file_data)
                            
                            # Hapus cellStyleXfs yang bermasalah
                            ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                            for elem in root.findall('.//cellStyleXfs', ns):
                                # Kosongkan atau set minimal
                                elem.clear()
                                elem.set('count', '0')
                            
                            # Rebuild XML
                            file_data = ET.tostring(root, encoding='utf-8')
                            print("    → Repaired styles.xml")
                        except:
                            print("    → Using original styles.xml")
                    
                    # Write file ke new ZIP
                    new_zip.writestr(filename, file_data)
            
            repaired_bytes.seek(0)
            return repaired_bytes
            
    except Exception as e:
        print(f"    → Repair failed: {e}")
        return None


def read_excel_with_repair(file_bytes, sheet_name=0, header=0, filename="file.xlsx"):
    """
    Read Excel file using auto-repair if that file corrupt
    """
    methods = [
        {
            'name': 'Standard openpyxl',
            'func': lambda fb: pd.read_excel(fb, sheet_name=sheet_name, header=header, engine='openpyxl')
        },
        {
            'name': 'Calamine engine (fast, ignore styles)',
            'func': lambda fb: pd.read_excel(fb, sheet_name=sheet_name, header=header, engine='calamine')
        },
        {
            'name': 'xlrd (for older formats)',
            'func': lambda fb: pd.read_excel(fb, sheet_name=sheet_name, header=header, engine='xlrd')
        }
    ]
    
    # Try standard methods first
    for method in methods:
        try:
            file_bytes.seek(0)
            df = method['func'](file_bytes)
            print(f"    ✓ Success with: {method['name']}")
            return df
        except Exception as e:
            print(f"    ✗ {method['name']}: {type(e).__name__}")
            continue
    
    # If all failed, try repair
    print("    → Attempting to repair file...")
    file_bytes.seek(0)
    repaired = repair_excel_styles(file_bytes)
    
    if repaired:
        try:
            df = pd.read_excel(repaired, sheet_name=sheet_name, header=header, engine='openpyxl')
            print(f"    ✓ Success after repair!")
            return df
        except Exception as e:
            print(f"    ✗ Still failed after repair: {e}")
    
    # Last resort: manual extraction dengan openpyxl
    print("    → Trying manual data extraction...")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill  # Import untuk default styles
        
        file_bytes.seek(0)
        
        # Save to temp file karena load_workbook butuh file path untuk repair mode
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(file_bytes.read())
            tmp_path = tmp.name
        
        try:
            # Repair on-the-fly dengan keep_links=False
            from openpyxl.reader.excel import load_workbook as openpyxl_load
            
            # Patch openpyxl untuk skip broken styles
            import openpyxl.styles.stylesheet
            original_init = openpyxl.styles.stylesheet.Stylesheet.__init__
            
            def patched_init(self, *args, **kwargs):
                try:
                    original_init(self, *args, **kwargs)
                except IndexError:
                    # If styles broken, use minimal defaults
                    print("    → Using default styles (stylesheet broken)")
                    self.named_styles = []
            
            openpyxl.styles.stylesheet.Stylesheet.__init__ = patched_init
            
            # Load workbook
            wb = load_workbook(tmp_path, data_only=True, keep_links=False)
            ws = wb.active if isinstance(sheet_name, int) and sheet_name == 0 else wb[sheet_name]
            
            # Extract data
            data = list(ws.values)
            
            # Restore original init
            openpyxl.styles.stylesheet.Stylesheet.__init__ = original_init
            
            # Create DataFrame
            if header is not None and len(data) > header:
                df = pd.DataFrame(data[header+1:], columns=data[header])
            else:
                df = pd.DataFrame(data)
            
            print(f"    ✓ Manual extraction successful!")
            return df
            
        finally:
            # Cleanup temp file
            try:
                os.unlink(tmp_path)
            except:
                pass
                
    except Exception as e:
        print(f"    ✗ Manual extraction failed: {type(e).__name__} - {str(e)[:100]}")
    
    raise Exception(f"Cannot read {filename} with any method")


def read_data(FOLDER_PATH, FILE_PATTERN, SHEET_NAME, HEADER, TOKEN, SITE_ID, DRIVE_ID, CSVDelimiter, NeedBackup, backup_folder_path):

    headers = {"Authorization": f"Bearer {TOKEN}"}

    folder_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/root:/{FOLDER_PATH}:/children"

    # Ambil semua file + recursive ke subfolder
    all_files = list_all_files(folder_url, FOLDER_PATH, headers, SITE_ID, DRIVE_ID)
    
    # Filter by regex
    matched_files = [f for f in all_files if re.fullmatch(FILE_PATTERN, f["name"])]

    if not matched_files:
        print("No files matched:", FILE_PATTERN)
        return None, None

    root_only = all(f["folder_name"] is None for f in matched_files)
    if root_only and len(matched_files) > 1:
        print(f"Multiple root-level files matched pattern. Taking one: {matched_files[0]['name']}")
        matched_files = [matched_files[0]]

    df_list = []
    TableName = None

    for f in matched_files:
        print(f"\n{'='*60}")
        print(f"Reading: {f['name']}")
        print('='*60)
        
        # CREATE BACKUP BEFORE READING
        if NeedBackup == 'Y':
            backup_result = backup_file_in_sharepoint(
                file_id=f["file_id"],
                file_name=f["name"],
                parent_folder_id=f.get("parent_folder_id"),
                headers=headers,
                SITE_ID=SITE_ID,
                DRIVE_ID=DRIVE_ID,
                backup_folder_path = backup_folder_path
            )
            
            if backup_result:
                print(f"  ✓ Backup completed: {backup_result['backup_name']}")
            else:
                print(f"  ⚠ Backup skipped or failed - continuing with read")
        
        TableName = f['name']
        r_file = requests.get(f["download_url"], headers=headers)

        if r_file.status_code != 200:
            print(f"  ✗ Failed to download (status {r_file.status_code})")
            continue

        file_bytes = BytesIO(r_file.content)
        file_size = len(r_file.content)
        print(f"  File size: {file_size:,} bytes")
        
        if file_size == 0:
            print(f"  ✗ File is empty (0 bytes)")
            continue

        try:
            # Excel
            if f["name"].lower().endswith(".xlsx"):
                df = read_excel_with_repair(
                    file_bytes, 
                    sheet_name=SHEET_NAME if SHEET_NAME else 0,
                    header=HEADER,
                    filename=f["name"]
                )
                
                if df.empty:
                    print(f"  ⚠ Warning: DataFrame is empty after reading")
                    # continue
                
                print(f"  → Result: {len(df)} rows × {len(df.columns)} columns")
            
            # Excel xls
            elif f["name"].lower().endswith(".xls"):
                df = read_excel_with_repair(
                    file_bytes, 
                    sheet_name=SHEET_NAME if SHEET_NAME else 0,
                    header=HEADER,
                    filename=f["name"]
                )
                
                if df.empty:
                    print(f"  ⚠ Warning: DataFrame is empty after reading")
                    # continue
                
                print(f"  → Result: {len(df)} rows × {len(df.columns)} columns")
                
            # CSV auto delimiter
            elif f["name"].lower().endswith(".csv"):
                text = file_bytes.getvalue().decode("utf-8", errors="ignore")

                valid_delimiters = ["comma", "semicolon", "tab", "pipe"]
                if CSVDelimiter not in valid_delimiters:
                    raise ValueError(f"Invalid CSVDelimiter: {CSVDelimiter}")

                # Mapping CSVDelimiter to actual delimiter
                if CSVDelimiter == "comma":
                    delimiter = ","
                elif CSVDelimiter == "semicolon":
                    delimiter = ";"
                elif CSVDelimiter == "tab":
                    delimiter = "\t"
                elif CSVDelimiter == "pipe":
                    delimiter = "|"

                df = pd.read_csv(
                    io.StringIO(text),
                    delimiter=delimiter,
                    header=HEADER
                )

                print(f"  → Result: {len(df)} rows × {len(df.columns)} columns")

            if len(matched_files) > 1:
                folder_name = f.get('folder_name')
                df['Source'] = folder_name if folder_name else 'Root'

            df_list.append(df)
            print(f"  ✓ Successfully added to dataset")

        except Exception as e:
            print(f"  ✗ SKIPPED: {type(e).__name__}")
            print(f"     {str(e)[:200]}")
            continue

    print(f"\n{'='*60}")
    if not df_list:
        print("❌ No files processed successfully")
        return None, None
    

    final_df = pd.concat(df_list, ignore_index=True)
    print(f"✓ SUCCESS! Combined dataset:")
    print(f"  - Total rows: {final_df.shape[0]:,}")
    print(f"  - Total columns: {final_df.shape[1]}")
    print(f"  - Files processed: {len(df_list)}")
    print('='*60)

    return final_df, TableName

# def get_columns(data):
    